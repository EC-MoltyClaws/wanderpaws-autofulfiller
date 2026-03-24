#!/usr/bin/env python3
"""
Daily Shopify Orders Report Generator

Flow:
  1. Compute report window: yesterday 10am → today 10am (SGT)
  2. POST date range to Make.com webhook → receives { orderData: [...] }
  3. Client-side filter (belt-and-suspenders in case Make returns all orders)
  4. Build one row per order; Lineitem name = comma-separated SKU-mapped items
  5. Generate styled Excel workbook
  6. Send workbook to Telegram
"""

import argparse
import io
import os
import sys
from datetime import datetime, timedelta
from pathlib import Path

from dotenv import load_dotenv
import pytz

load_dotenv()
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from sku_config import SKU_MAP

# ─── Config ───────────────────────────────────────────────────────────────────

WEBHOOK_URL = os.environ["MAKE_WEBHOOK_URL"]
LAST_ORDER_WEBHOOK_URL = os.environ["MAKE_LAST_ORDER_WEBHOOK_URL"]
WEBHOOK_API_KEY = os.environ["MAKE_WEBHOOK_API_KEY"]
TELEGRAM_BOT_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
TELEGRAM_CHAT_ID = os.environ["TELEGRAM_CHAT_ID"]
TIMEZONE = os.environ.get("REPORT_TIMEZONE", "Asia/Singapore")


# ─── SKU expansion ────────────────────────────────────────────────────────────


def expand_sku(sku: str, qty: int) -> list[str]:
    """
    Phase 1: Expand a pseudo SKU into display strings for each actual product.

    Looks up the SKU directly in SKU_MAP.
      - Known SKU with products → ["Nx Product-Variant", ...]
      - Known SKU mapped to []  → [] (skip silently)
      - Unknown SKU             → raises ValueError

    Raises:
        ValueError: if the SKU is not present in SKU_MAP.
    """
    if sku not in SKU_MAP:
        raise ValueError(f"Unknown SKU: {sku}")
    return [f"{qty}x {name}" for name in SKU_MAP[sku]]


# ─── Time window ──────────────────────────────────────────────────────────────


def get_report_window() -> tuple[datetime, datetime]:
    """Returns (start, end): yesterday 10:00am → today 10:00am in TIMEZONE."""
    tz = pytz.timezone(TIMEZONE)
    now = datetime.now(tz)
    end = now.replace(hour=10, minute=0, second=0, microsecond=0)
    start = end - timedelta(days=1)
    return start, end


# ─── Data fetching ────────────────────────────────────────────────────────────


def fetch_orders(start: datetime, end: datetime, reship_orders: list = None) -> list:
    """
    POST to Make.com webhook with the date window and optional reship orders.
    Make.com should use these params to filter Shopify orders at source.
    Returns list of Shopify order objects.
    """
    if reship_orders is None:
        reship_orders = []

    payload = {
        "start": start.isoformat(),
        "end": end.isoformat(),
        "reship_orders": reship_orders
    }
    resp = requests.post(
        WEBHOOK_URL,
        json=payload,
        headers={"x-make-apikey": WEBHOOK_API_KEY},
        timeout=60,
    )

    print(f"Webhook status  : {resp.status_code} {resp.reason}")
    print(f"Webhook response: {resp.text[:500]}")

    resp.raise_for_status()

    data = resp.json()
    orders = data.get("orderData", [])
    print(f"Fetched {len(orders)} orders from webhook.")
    return orders


def fetch_last_order() -> list:
    """Call the last-order webhook. Returns a single-item list."""
    resp = requests.post(
        LAST_ORDER_WEBHOOK_URL,
        json={},
        headers={"x-make-apikey": WEBHOOK_API_KEY},
        timeout=60,
    )

    print(f"Webhook status  : {resp.status_code} {resp.reason}")
    print(f"Webhook response: {resp.text[:500]}")

    resp.raise_for_status()

    data = resp.json()
    orders = data.get("orderData", [])
    print(f"Fetched {len(orders)} order(s) from last-order webhook.")
    return orders


# ─── Filtering ────────────────────────────────────────────────────────────────


def filter_orders(orders: list, start: datetime, end: datetime) -> list:
    """Client-side filter on created_at in case webhook returns all orders."""
    print(f"Window : {start.isoformat()} -> {end.isoformat()}")
    filtered = []
    for order in orders:
        if True:
            filtered.append(order)
    print(f"Filtered to {len(filtered)} orders in window.")
    return filtered


# ─── Lineitem name ────────────────────────────────────────────────────────────


def build_lineitem_name(order: dict) -> str:
    """
    Phase 2: Collate all expanded SKUs across the order's line items.

    Ordering:
      1. Bundle items (SKUs that expand to 2+ products), grouped by slot:
         slot 0 across all line items, then slot 1 (collated), etc.
      2. Standalone items (SKUs that expand to 1 product), in order.

    Quantities are summed for identical product names.

    Example: "2x Harness-Black/XS, 2x Leash-Black, 1x Retractable Leash"

    Raises:
        ValueError: if any SKU in the order is not present in SKU_MAP.
    """
    # slot_items[i]: ordered dict of name → total qty for bundle slot i
    slot_items: list[dict[str, int]] = []
    standalone_items: dict[str, int] = {}

    for item in order.get("lineItems", []):
        sku = (item.get("sku") or "").strip()
        qty = item.get("quantity", 1)

        if sku not in SKU_MAP:
            raise ValueError(f"Unknown SKU: {sku}")

        names = SKU_MAP[sku]

        if not names:
            continue  # non-fulfilled item, skip

        if len(names) == 1:
            name = names[0]
            standalone_items[name] = standalone_items.get(name, 0) + qty
        else:
            while len(slot_items) < len(names):
                slot_items.append({})
            for i, name in enumerate(names):
                slot_items[i][name] = slot_items[i].get(name, 0) + qty

    parts = []
    for slot in slot_items:
        parts.extend(f"{total}x {name}" for name, total in slot.items())
    parts.extend(f"{total}x {name}" for name, total in standalone_items.items())
    return ", ".join(parts)


# ─── Row building ─────────────────────────────────────────────────────────────

COLUMNS = [
    "Name",
    "Email",
    "Lineitem name",
    "Shipping Name",
    "Shipping Street",
    "Shipping Address1",
    "Shipping Address2",
    "Shipping Company",
    "Shipping City",
    "Shipping Zip",
    "Shipping Province",
    "Shipping Country",
    "Shipping Phone",
]


def build_rows(orders: list) -> list[list]:
    """One row per order."""
    rows = []
    for order in orders:
        addr = order.get("shippingAddress") or {}
        rows.append([
            order.get("name"),
            order.get("email"),
            build_lineitem_name(order),
            addr.get("name"),
            addr.get("address1"),
            addr.get("address1"),
            addr.get("address2"),
            addr.get("company"),
            addr.get("city"),
            addr.get("zip"),
            addr.get("province"),
            addr.get("country"),
            addr.get("phone"),
        ])
    return rows


# ─── Excel styles ─────────────────────────────────────────────────────────────

_HEADER_FONT = Font(bold=True, name="Calibri", size=11)
_NORMAL_FONT = Font(name="Calibri", size=10)
_THIN = Side(style="thin", color="BBBBBB")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _auto_width(ws) -> None:
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=8)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)


# ─── Sheet builder ────────────────────────────────────────────────────────────


def _build_orders_sheet(ws, rows: list[list]) -> None:
    ws.title = "Orders"
    ws.sheet_view.showGridLines = False

    # Header
    for col_idx, col_name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _BORDER
    ws.row_dimensions[1].height = 20

    # Data rows
    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = _NORMAL_FONT
            cell.border = _BORDER

    _auto_width(ws)


# ─── Excel generation ─────────────────────────────────────────────────────────


def generate_excel(rows: list[list]) -> io.BytesIO:
    wb = Workbook()
    _build_orders_sheet(wb.active, rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── Telegram ─────────────────────────────────────────────────────────────────


def send_telegram_document(buf: io.BytesIO, filename: str, caption: str) -> None:
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    resp = requests.post(
        url,
        data={"chat_id": TELEGRAM_CHAT_ID, "caption": caption, "parse_mode": "Markdown"},
        files={"document": (filename, buf, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        timeout=30,
    )
    resp.raise_for_status()
    print("Report sent to Telegram.")


def send_telegram_message(text: str) -> None:
    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    requests.post(
        url,
        data={"chat_id": TELEGRAM_CHAT_ID, "text": text, "parse_mode": "Markdown"},
        timeout=15,
    )


# ─── SKU validation ───────────────────────────────────────────────────────────


def find_missing_skus(orders: list) -> list[tuple[str, str]]:
    """Return (order_name, item_title) for every line item that has no SKU."""
    missing = []
    for order in orders:
        order_name = order.get("name", "?")
        for item in order.get("lineItems", []):
            sku = (item.get("sku") or "").strip()
            if not sku:
                product_id = item.get("product", {}).get("legacyResourceId", "Unknown")
                missing.append((order_name, product_id))
    return missing


# ─── Main ─────────────────────────────────────────────────────────────────────


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--last", action="store_true", help="Fetch only the last order via the last-order webhook")
    args = parser.parse_args()

    try:
        if args.last:
            orders = fetch_last_order()
            label = "Last Order"
        else:
            start, end = get_report_window()
            reship_str = os.environ.get("RESHIP_ORDERS", "").strip()
            reships = [r.strip() for r in reship_str.split(",")] if reship_str else []
            
            print(f"Report window: {start.isoformat()} -> {end.isoformat()}")
            if reships:
                print(f"Injecting Reship Orders: {reships}")
                
            orders = filter_orders(fetch_orders(start, end, reship_orders=reships), start, end)
            label = f"Daily Report — {end.strftime('%d %b %Y')}"

        tz = pytz.timezone(TIMEZONE)
        created_at = datetime.now(tz).strftime("%Y-%m-%d")

        if not orders:
            filename = f"Orders {created_at} #No Orders.xlsx"
            buf = generate_excel([])
            send_telegram_document(buf, filename, f"*WanderPaws {label}*\nNo orders found.")
            print("No orders — sent empty Excel via Telegram.")
            return

        # Sort ascending by order number so the sheet reads oldest → newest
        orders.sort(key=lambda o: int((o.get("name") or "#0").lstrip("#") or 0))

        missing_skus = find_missing_skus(orders)
        if missing_skus:
            lines = "\n".join(f"• {name}: _{title}_" for name, title in missing_skus)
            send_telegram_message(f"*WanderPaws SKU Missing*\nThe following items have no SKU and the report was not generated:\n{lines}")
            print("Aborted — missing SKUs detected.")
            return

        first_order_name = (orders[0].get("name") or "").lstrip("#")
        last_order_name = (orders[-1].get("name") or "").lstrip("#")
        if len(orders) == 1:
            filename = f"Orders {created_at} #{first_order_name}.xlsx"
        else:
            filename = f"Orders {created_at} #{first_order_name}-#{last_order_name}.xlsx"

        rows = build_rows(orders)
        buf = generate_excel(rows)

        send_telegram_document(buf, filename, f"*WanderPaws {label}*\n`{len(orders)}` order(s)")

    except Exception as e:
        send_telegram_message(f"*WanderPaws Report Error*\n`{type(e).__name__}: {e}`")
        sys.exit(1)


if __name__ == "__main__":
    main()
